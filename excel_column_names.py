import os

try:
	import openpyxl
except:
	print("could not find openpyxl module")




def getFilesInDirectory(directory):
	fileNames = []
	excelFile_Paths = []

	current_path = os.getcwd()
	for file_in_directory in os.listdir(os.path.join(current_path, directory)):
		if file_in_directory.endswith(".xlsx"):
			excelFile_Paths.append(os.path.join(current_path, file_in_directory))
			fileNames.append(file_in_directory)

	return fileNames, excelFile_Paths

def convert_color_code(color_index):
    if isinstance(color_index, str):
        return color_index
    if isinstance(color_index, int):
        if color_index <= 63:
            return COLOR_INDEX[color_index]
        else:
            return "System background"


def titleRowsOfSheet(sheet, sheetname):
	# row_count = sheet.max_row
	# col_count = sheet.max_column
	mergedRanges = sheet.merged_cells.ranges
	title_rows = []

	# values update in loop
	column_max = 1
	column_min = sheet.max_column

	for row in sheet.iter_rows():
		all_filled = True
		all_empty = True

		for cell in row:
			if type(cell).__name__ == "MergedCell":
				first_cell = [cell_in_merge for cell_in_merge in sheet.merged_cells.ranges if cell.coordinate in cell_in_merge][0].start_cell
				merged_cell_color_index = first_cell.fill.end_color.index
				color_value = convert_color_code(merged_cell_color_index)
			else:
				color_index = cell.fill.end_color.index
				color_value = convert_color_code(color_index)

			if color_value != "00000000":
				# print("PASS "+ cell.coordinate)
				if cell.column > column_max: 
					column_max = cell.column 
				if cell.column < column_min:
					column_min = cell.column
				all_empty = False

			else:
				# print("FAIL "+ cell.coordinate)

				if cell.value != None:
					all_filled = False
					break;
				

		if all_filled and not all_empty:
			title_rows.append(row[0].row)  # add row numbers

	print(f"sheet: {sheetname}, min: {column_min}, max: {column_max}, title rows: {title_rows}")

	return column_min, column_max, title_rows

def columnNamesOfSheet(sheet, title_rows, min_column, max_column, sheetname):
	title_values = []
	column_count = sheet.max_column
	for col in range(min_column, max_column+1):
		column_values = []

		for row in range(title_rows[0], title_rows[-1] + 1):
			cell = sheet.cell(row=row, column=col)
	
			if type(cell).__name__ == "MergedCell":
				cell_value = [cell_in_merge for cell_in_merge in sheet.merged_cells.ranges if cell.coordinate in cell_in_merge][0].start_cell.value
			else:
				cell_value = cell.value
			
			if cell_value == None:
				cell_value = ""

			cell_value = str(cell_value)

			if (cell_value not in column_values):
				column_values.append(cell_value)
	
		title_values.append(" ".join(column_values))

	print(f"{sheetname}: {title_values}")
	return title_values



def main():
	fileNames, excelFile_Paths = getFilesInDirectory(directory)
	current_path = os.getcwd()

	for fileIndex in range(0, len(excelFile_Paths)):
		excelFile = openpyxl.load_workbook(excelFile_Paths[fileIndex])

		folder_name = fileNames[fileIndex].replace(".xlsx", "")
		folder_path = os.path.join(current_path, folder_name)

		try:
			os.mkdir(folder_path)
		except (FileExistsError):
			print(f"{folder_name} already exists")


		if set_sheetname_part:
			for sheetname in excelFile.sheetnames:
				if set_sheetname_part not in sheetname:
					continue

				sheet = excelFile[sheetname]
				column_min, column_max, title_rows = titleRowsOfSheet(sheet, sheetname)
				columnNames = columnNamesOfSheet(sheet, title_rows, column_min, column_max, sheetname)
	
				outputFile_path = os.path.join(folder_path, sheetname + ".csv")
	
				with open(outputFile_path, "w") as outputFile:
					for i in range(0, len(columnNames) - 1):
						outputFile.write(columnNames[i] + ',')
					outputFile.write(columnNames[len(columnNames) - 1] + '\n')
	
					for row in sheet.iter_rows(min_row=title_rows[-1]+1):
						for cell in row:
							if cell.value != None:
								cell_value = str(cell.value)
								cell_value = cell_value.replace("\n", " ")
								cell_value = cell_value.replace(",", ";")
								
							else:
								cell_value = ""

							outputFile.write(cell_value + ',')

						outputFile.write('\n')

		else:
			for sheetname in excelFile.sheetnames:
				sheet = excelFile[sheetname]
				column_min, column_max, title_rows = titleRowsOfSheet(sheet, sheetname)
				columnNames = columnNamesOfSheet(sheet, title_rows, column_min, column_max, sheetname)
				outputFile_path = os.path.join(folder_path, sheetname + ".csv")
	
				with open(outputFile_path, "w") as outputFile:
					for i in range(0, len(columnNames) - 1):
						outputFile.write(columnNames[i] + ',')

					outputFile.write(columnNames[len(columnNames) - 1] + '\n')
	
					for row in sheet.iter_rows(min_row=title_rows[-1]+1):
						for cell in row:
							if cell.value != None:
								cell_value = str(cell.value)
								cell_value = cell_value.replace("\n", " ")
								cell_value = cell_value.replace(",", ";")
								
							else:
								cell_value = ""
							outputFile.write(cell_value + ',')

						outputFile.write('\n')


if __name__ == "__main__":
	set_sheetname_part = "IPX"
	directory = "" # input("Folder: ")
	main()